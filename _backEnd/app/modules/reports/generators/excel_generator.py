"""
Mitrufely Web — Excel Generator (Fase 7)
Genera reportes tabulares en .xlsx con `openpyxl` (multiplataforma).

Aplica estilos institucionales (cabecera vino, KPIs, totales) y produce un
workbook en memoria listo para descargar por el cliente.
"""

from __future__ import annotations

import io
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.modules.reports.schemas import (
    ReporteCatalogoResponse,
    ReporteFidelizacionResponse,
    ReporteInventarioResponse,
    ReportePedidosResponse,
    ReporteUsuariosResponse,
    ReporteVentasResponse,
)

# ── Estilos institucionales ───────────────────────────────────────────────────

VINO = "5C0F1B"
GRIS_CLARO = "FAF8F5"
NARANJA = "FF7A45"
BLANCO = "FFFFFF"

_FONT_TITULO = Font(name="Calibri", size=16, bold=True, color=VINO)
_FONT_SUB = Font(name="Calibri", size=10, color="2A1115")
_FONT_HEADER = Font(name="Calibri", size=11, bold=True, color=BLANCO)
_FONT_CELL = Font(name="Calibri", size=10, color="2A1115")
_FONT_KPI_LBL = Font(name="Calibri", size=9, color="888888")
_FONT_KPI_VAL = Font(name="Calibri", size=14, bold=True, color=VINO)

_FILL_HEADER = PatternFill("solid", fgColor=VINO)
_FILL_ROW_ALT = PatternFill("solid", fgColor=GRIS_CLARO)
_FILL_KPI = PatternFill("solid", fgColor=GRIS_CLARO)

_BORDER = Border(
    left=Side(style="thin", color="E7E5E4"),
    right=Side(style="thin", color="E7E5E4"),
    top=Side(style="thin", color="E7E5E4"),
    bottom=Side(style="thin", color="E7E5E4"),
)

_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
_ALIGN_LEFT = Alignment(horizontal="left", vertical="center")


def _money(v) -> str:
    try:
        return f"S/ {Decimal(str(v)).quantize(Decimal('0.01'))}"
    except Exception:
        return str(v)


def _wb_vacio(titulo: str) -> Workbook:
    wb = Workbook()
    wb.active.title = titulo[:31]
    return wb


def _escribir_encabezado(ws, titulo: str, subtitulo: str, n_cols: int) -> int:
    """Escribe título + subtítulo y devuelve la fila siguiente libre."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    c = ws.cell(row=1, column=1, value=titulo)
    c.font = _FONT_TITULO
    c.alignment = _ALIGN_LEFT

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    c2 = ws.cell(row=2, column=1, value=subtitulo)
    c2.font = _FONT_SUB
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 16
    return 4  # deja una fila en blanco


def _escribir_kpis(ws, row: int, kpis: list[tuple[str, str]], n_cols: int) -> int:
    """Escribe una tira de KPIs a partir de `row`. Devuelve la fila siguiente."""
    if not kpis:
        return row + 1
    col = 1
    bloque = max(1, n_cols // max(len(kpis), 1))
    for label, value in kpis:
        # etiqueta
        ws.cell(row=row, column=col, value=label.upper()).font = _FONT_KPI_LBL
        ws.cell(row=row, column=col).alignment = _ALIGN_CENTER
        ws.cell(row=row, column=col).fill = _FILL_KPI
        # valor
        ws.cell(row=row + 1, column=col, value=value).font = _FONT_KPI_VAL
        ws.cell(row=row + 1, column=col).alignment = _ALIGN_CENTER
        ws.cell(row=row + 1, column=col).fill = _FILL_KPI
        for cc in range(col, col + bloque):
            ws.cell(row=row, column=cc).fill = _FILL_KPI
            ws.cell(row=row + 1, column=cc).fill = _FILL_KPI
        col += bloque
    ws.row_dimensions[row].height = 14
    ws.row_dimensions[row + 1].height = 22
    return row + 3


def _escribir_tabla(ws, row: int, headers: list[str], filas: list[list], widths: list[int]) -> int:
    # cabecera
    for j, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=j, value=h)
        c.font = _FONT_HEADER
        c.fill = _FILL_HEADER
        c.alignment = _ALIGN_CENTER
        c.border = _BORDER
    ws.row_dimensions[row].height = 22

    # filas
    for i, fila in enumerate(filas, start=row + 1):
        alt = (i - row) % 2 == 0
        for j, val in enumerate(fila, start=1):
            c = ws.cell(row=i, column=j, value=val)
            c.font = _FONT_CELL
            c.border = _BORDER
            c.alignment = _ALIGN_LEFT
            if alt:
                c.fill = _FILL_ROW_ALT

    # anchos
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w

    return row + 1 + len(filas)


# ── Exportador público ────────────────────────────────────────────────────────


def exportar_reporte_a_excel(reporte_tipo: str, data) -> bytes:
    """Serializa cualquier reporte (DTO Response) a .xlsx. Retorna los bytes."""
    if reporte_tipo == "ventas":
        wb = _xlsx_ventas(data)
    elif reporte_tipo == "pedidos":
        wb = _xlsx_pedidos(data)
    elif reporte_tipo == "catalogo":
        wb = _xlsx_catalogo(data)
    elif reporte_tipo == "inventario":
        wb = _xlsx_inventario(data)
    elif reporte_tipo == "usuarios":
        wb = _xlsx_usuarios(data)
    elif reporte_tipo == "fidelizacion":
        wb = _xlsx_fidelizacion(data)
    else:
        raise ValueError(f"Tipo de reporte no soportado: {reporte_tipo}")

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ── Implementaciones por reporte ──────────────────────────────────────────────


def _xlsx_ventas(r: ReporteVentasResponse) -> Workbook:
    wb = _wb_vacio("Ventas")
    ws = wb.active
    n_cols = 9
    row = _escribir_encabezado(
        ws,
        "Reporte de Rendimiento de Ventas",
        "Comportamiento económico · productos más vendidos · métodos de pago.",
        n_cols,
    )
    row = _escribir_kpis(
        ws, row,
        [
            ("Total ventas", _money(r.total_ventas)),
            ("Pedidos", str(r.cantidad_pedidos)),
            ("Ticket promedio", _money(r.ticket_promedio)),
        ],
        n_cols,
    )
    headers = ["ID", "Fecha", "Cliente", "Estado", "Pago", "Base", "IGV", "Total", "M. Pago"]
    filas = [
        [
            it.id_venta,
            it.fecha_venta.strftime("%d/%m/%Y %H:%M"),
            it.cliente, it.estado, it.estado_pago,
            _money(it.base_imponible), _money(it.igv), _money(it.total),
            it.metodo_pago or "—",
        ]
        for it in r.items
    ]
    _escribir_tabla(
        ws, row, headers, filas,
        [8, 18, 28, 14, 12, 14, 12, 14, 14],
    )
    return wb


def _xlsx_pedidos(r: ReportePedidosResponse) -> Workbook:
    wb = _wb_vacio("Pedidos")
    ws = wb.active
    n_cols = 7
    row = _escribir_encabezado(
        ws,
        "Reporte de Seguimiento de Pedidos",
        "Estado de pedidos · pendientes / completados / entregados.",
        n_cols,
    )
    resumen = " · ".join(f"{k}: {v}" for k, v in r.por_estado.items()) or "Sin datos"
    row = _escribir_kpis(ws, row, [("Distribución", resumen[:40]), ("Total", str(r.total_pedidos))], n_cols)

    headers = ["ID", "Cliente", "Estado", "Pago", "Fecha venta", "Entregado", "Total"]
    filas = [
        [
            it.id_venta, it.cliente, it.estado, it.estado_pago,
            it.fecha_venta.strftime("%d/%m/%Y %H:%M"),
            (it.delivery_completed_at.strftime("%d/%m/%Y") if it.delivery_completed_at else "—"),
            _money(it.total_final),
        ]
        for it in r.items
    ]
    _escribir_tabla(ws, row, headers, filas, [8, 28, 14, 12, 18, 14, 14])
    return wb


def _xlsx_catalogo(r: ReporteCatalogoResponse) -> Workbook:
    wb = _wb_vacio("Catálogo")
    ws = wb.active
    n_cols = 6
    row = _escribir_encabezado(
        ws,
        "Reporte de Catálogo Comercial",
        "Productos registrados · estado · detección de inactivos.",
        n_cols,
    )
    row = _escribir_kpis(
        ws, row,
        [("Total", str(r.total_productos)), ("Activos", str(r.productos_activos)), ("Inactivos", str(r.productos_inactivos))],
        n_cols,
    )
    headers = ["ID", "Nombre", "Categoría", "Precio", "Stock", "Estado"]
    filas = [
        [
            it.id_producto, it.nombre, it.categoria or "—",
            _money(it.precio), it.stock_actual,
            "Activo" if it.estado else "Inactivo",
        ]
        for it in r.items
    ]
    _escribir_tabla(ws, row, headers, filas, [8, 36, 22, 14, 10, 12])
    return wb


def _xlsx_inventario(r: ReporteInventarioResponse) -> Workbook:
    wb = _wb_vacio("Inventario")
    ws = wb.active
    n_cols = 7
    row = _escribir_encabezado(
        ws,
        "Reporte de Control de Inventario",
        "Stock disponible · productos agotados o con bajo stock · valorización.",
        n_cols,
    )
    row = _escribir_kpis(
        ws, row,
        [
            ("Productos", str(r.total_productos)),
            ("Bajo stock", str(r.productos_bajo_stock)),
            ("Agotados", str(r.productos_agotados)),
            ("Valor inv.", _money(r.valor_inventario)),
        ],
        n_cols,
    )
    headers = ["ID", "Nombre", "Categoría", "Stock", "Mínimo", "Estado stock", "Valorización"]
    filas = [
        [
            it.id_producto, it.nombre, it.categoria or "—",
            it.stock_actual, it.stock_minimo, it.estado_stock, _money(it.valorizacion),
        ]
        for it in r.items
    ]
    _escribir_tabla(ws, row, headers, filas, [8, 32, 22, 10, 10, 14, 16])
    return wb


def _xlsx_usuarios(r: ReporteUsuariosResponse) -> Workbook:
    wb = _wb_vacio("Usuarios")
    ws = wb.active
    n_cols = 7
    row = _escribir_encabezado(
        ws,
        "Reporte de Gestión de Usuarios",
        "Personas registradas · roles · estado de cuentas · actividad.",
        n_cols,
    )
    row = _escribir_kpis(
        ws, row,
        [("Total", str(r.total_usuarios)), ("Activos", str(r.activos)), ("Inactivos", str(r.inactivos))],
        n_cols,
    )
    headers = ["ID", "Nombres", "Apellidos", "Email", "Rol", "Estado", "Auth"]
    filas = [
        [
            it.id_usuario, it.nombres, it.apellidos, it.email, it.rol,
            "Activo" if it.estado else "Inactivo", it.auth_provider,
        ]
        for it in r.items
    ]
    _escribir_tabla(ws, row, headers, filas, [8, 22, 22, 32, 12, 12, 12])
    return wb


def _xlsx_fidelizacion(r: ReporteFidelizacionResponse) -> Workbook:
    wb = _wb_vacio("Fidelización")
    ws = wb.active
    n_cols = 5
    row = _escribir_encabezado(
        ws,
        "Reporte de Fidelización SweetCoins / CriptoTrufa",
        "Puntos acumulados · utilizados · próximos a vencer.",
        n_cols,
    )
    row = _escribir_kpis(
        ws, row,
        [("Clientes", str(r.total_clientes)), ("Puntos circulación", str(r.puntos_circulacion))],
        n_cols,
    )
    headers = ["ID", "Cliente", "Email", "Saldo puntos", "Puntos usados"]
    filas = [
        [it.id_cliente, it.cliente, it.email, it.saldo_puntos, it.puntos_usados]
        for it in r.items
    ]
    _escribir_tabla(ws, row, headers, filas, [8, 32, 32, 16, 16])
    return wb
